import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def crear_excel_liga():
    wb = openpyxl.Workbook()
    
    # Colores y Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def aplicar_estilos_header(ws, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            # Ajustar ancho de columna
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    def aplicar_bordes_y_centro(ws, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border

    # ==========================
    # 1. Hoja "Equipos"
    # ==========================
    ws_equipos = wb.active
    ws_equipos.title = "Equipos"
    ws_equipos.append(["Equipo (Amigo)"])
    equipos = ["Mariopo", "Dynn", "diego", "Eduardo", "Khalvito", "Manko"]
    for eq in equipos:
        ws_equipos.append([eq])
    
    aplicar_estilos_header(ws_equipos, 1)
    aplicar_bordes_y_centro(ws_equipos, 2, 7, 1, 1)

    # ==========================
    # 2. Hoja "Resultados"
    # ==========================
    ws_resultados = wb.create_sheet(title="Resultados")
    headers_res = ["Jornada", "Equipo Local", "Goles Local", "Goles Visitante", "Equipo Visitante"]
    ws_resultados.append(headers_res)
    aplicar_estilos_header(ws_resultados, 5)
    ws_resultados.column_dimensions['B'].width = 20
    ws_resultados.column_dimensions['E'].width = 20

    # Validación de datos para listas desplegables
    dv = DataValidation(type="list", formula1="=Equipos!$A$2:$A$7", allow_blank=True)
    ws_resultados.add_data_validation(dv)
    # Aplicar a las filas 2 a 100 en las columnas B (Local) y E (Visitante)
    dv.add("B2:B100")
    dv.add("E2:E100")

    # ==========================
    # 3. Hoja Oculta "Calculos_Liga"
    # ==========================
    ws_calc = wb.create_sheet(title="Calculos_Liga")
    headers_calc = ["Equipo", "PJ", "PG", "PE", "PP", "GF", "GC", "DG", "Pts", "Score", "Rank"]
    ws_calc.append(headers_calc)
    
    for i in range(2, 8):
        # A: Equipo
        ws_calc[f"A{i}"] = f"=Equipos!A{i}"
        
        # B: PJ
        ws_calc[f"B{i}"] = f"=SUMPRODUCT((Resultados!$B$2:$B$100=A{i})*(Resultados!$C$2:$C$100<>\"\")) + SUMPRODUCT((Resultados!$E$2:$E$100=A{i})*(Resultados!$D$2:$D$100<>\"\"))"
        
        # C: PG
        ws_calc[f"C{i}"] = f"=SUMPRODUCT((Resultados!$B$2:$B$100=A{i})*(Resultados!$C$2:$C$100>Resultados!$D$2:$D$100)*(Resultados!$C$2:$C$100<>\"\")) + SUMPRODUCT((Resultados!$E$2:$E$100=A{i})*(Resultados!$D$2:$D$100>Resultados!$C$2:$C$100)*(Resultados!$C$2:$C$100<>\"\"))"
        
        # D: PE
        ws_calc[f"D{i}"] = f"=SUMPRODUCT((Resultados!$B$2:$B$100=A{i})*(Resultados!$C$2:$C$100=Resultados!$D$2:$D$100)*(Resultados!$C$2:$C$100<>\"\")) + SUMPRODUCT((Resultados!$E$2:$E$100=A{i})*(Resultados!$D$2:$D$100=Resultados!$C$2:$C$100)*(Resultados!$D$2:$D$100<>\"\"))"
        
        # E: PP
        ws_calc[f"E{i}"] = f"=SUMPRODUCT((Resultados!$B$2:$B$100=A{i})*(Resultados!$C$2:$C$100<Resultados!$D$2:$D$100)*(Resultados!$C$2:$C$100<>\"\")) + SUMPRODUCT((Resultados!$E$2:$E$100=A{i})*(Resultados!$D$2:$D$100<Resultados!$C$2:$C$100)*(Resultados!$C$2:$C$100<>\"\"))"
        
        # F: GF
        ws_calc[f"F{i}"] = f"=SUMIF(Resultados!$B$2:$B$100, A{i}, Resultados!$C$2:$C$100) + SUMIF(Resultados!$E$2:$E$100, A{i}, Resultados!$D$2:$D$100)"
        
        # G: GC
        ws_calc[f"G{i}"] = f"=SUMIF(Resultados!$B$2:$B$100, A{i}, Resultados!$D$2:$D$100) + SUMIF(Resultados!$E$2:$E$100, A{i}, Resultados!$C$2:$C$100)"
        
        # H: DG
        ws_calc[f"H{i}"] = f"=F{i}-G{i}"
        
        # I: Pts
        ws_calc[f"I{i}"] = f"=C{i}*3 + D{i}*1"
        
        # J: Score
        ws_calc[f"J{i}"] = f"=I{i}*10000 + H{i}*100 + F{i}"
        
        # K: Rank
        ws_calc[f"K{i}"] = f"=RANK.EQ(J{i}, $J$2:$J$7) + COUNTIF($J$2:J{i}, J{i}) - 1"

    ws_calc.sheet_state = 'hidden' # Ocultamos la hoja de cálculos para que quede limpio

    # ==========================
    # 4. Hoja "Posiciones"
    # ==========================
    ws_posiciones = wb.create_sheet(title="Posiciones", index=1) # La ponemos segunda
    headers_pos = ["Posición", "Equipo", "PJ", "PG", "PE", "PP", "GF", "GC", "DG", "Pts"]
    ws_posiciones.append(headers_pos)
    aplicar_estilos_header(ws_posiciones, 10)
    ws_posiciones.column_dimensions['B'].width = 20

    for i in range(2, 8):
        # A: Posición 1 a 6
        ws_posiciones[f"A{i}"] = i - 1
        
        # Buscar el equipo y estadísticas según el Rank
        for col_idx, col_letter in enumerate(['B','C','D','E','F','G','H','I','J'], start=1): # Columnas de B a J en Posiciones corresponden a A a I en Calculos
            calc_col_letter = openpyxl.utils.get_column_letter(col_idx)
            # INDEX(Calculos_Liga!A$2:A$7, MATCH($A2, Calculos_Liga!$K$2:$K$7, 0))
            ws_posiciones[f"{col_letter}{i}"] = f"=INDEX(Calculos_Liga!{calc_col_letter}$2:{calc_col_letter}$7, MATCH($A{i}, Calculos_Liga!$K$2:$K$7, 0))"

    aplicar_bordes_y_centro(ws_posiciones, 2, 7, 1, 10)

    # Guardar el archivo
    filename = "Liga_FC26.xlsx"
    wb.save(filename)
    print(f"Archivo '{filename}' generado con éxito.")

if __name__ == "__main__":
    crear_excel_liga()
